"""
Gradio UI for Invoize.

Why Gradio instead of React/Next.js?
- 10x faster to build for a portfolio project (minutes, not days)
- Built-in file upload, JSON display, and download components
- Python-native — no JS build toolchain, no npm, no webpack
- Still looks professional enough for a demo
- Lets us focus time on the DIFFERENTIATOR (validation layer)
  instead of frontend plumbing

The UI calls the FastAPI backend over HTTP, keeping frontend and backend
cleanly separated — same architecture as a "real" app, just with Gradio
as the presentation layer instead of React.
"""

import sys
import os

# Fix sys.path to prevent 'frontend/app.py' from shadowing the backend 'app' package
root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if root_dir not in sys.path:
    sys.path.insert(0, root_dir)
current_dir = os.path.dirname(os.path.abspath(__file__))
while current_dir in sys.path:
    sys.path.remove(current_dir)

import json
import time
import tempfile
import csv
import httpx
import gradio as gr
import io
from PIL import Image

API_BASE = os.getenv("API_BASE", f"http://127.0.0.1:{os.getenv('PORT', '8000')}")


def preview_file(file):
    """Generate a PIL Image preview of an uploaded image or the first page of a PDF."""
    if not file:
        return None
    try:
        # Resolve actual file path from string, object, or dict
        if isinstance(file, list):
            if not file:
                return None
            file = file[0]

        if isinstance(file, str):
            file_path = file
        elif isinstance(file, dict):
            file_path = file.get("path") or file.get("name") or str(file)
        else:
            file_path = getattr(file, "path", None) or getattr(file, "name", None) or str(file)

        if not file_path or not os.path.exists(file_path):
            return None

        ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""
        if ext in ("csv", "json", "txt"):
            return None

        if ext == "pdf":
            import fitz
            doc = fitz.open(file_path)
            if doc.page_count > 0:
                page = doc[0]
                # Render page to a pixmap (DPI 150 is plenty for local preview)
                pix = page.get_pixmap(matrix=fitz.Matrix(150 / 72.0, 150 / 72.0))
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                doc.close()
                return img
            doc.close()
            return None
        else:
            return file_path
    except Exception as e:
        print(f"Error rendering preview: {e}")
        return None


def format_single_receipt_html(receipt: dict, filename: str = None) -> str:
    from app.schemas import get_currency_symbol
    sym = get_currency_symbol(receipt.get("currency"))
    
    # Header Details
    vendor = receipt.get("vendor_name", "Unknown Store")
    address = receipt.get("vendor_address") or ""
    date_val = receipt.get("date") or "—"
    time_val = receipt.get("time") or "—"
    currency = receipt.get("currency") or "INR"
    pay_method = receipt.get("payment_method") or "—"
    
    html = []
    html.append(f'<div class="receipt-card">')
    
    # Header block
    html.append(f'  <div class="receipt-card-header">')
    if filename:
        html.append(f'    <div class="receipt-card-subtitle" style="font-size: 11px; opacity: 0.6; font-family: monospace;">File: {filename}</div>')
    html.append(f'    <h2 class="receipt-card-title">{vendor}</h2>')
    if address:
        html.append(f'    <div class="receipt-card-subtitle">{address}</div>')
    html.append(f'  </div>')
    
    # Metadata Grid
    html.append(f'  <div class="receipt-meta-grid">')
    html.append(f'    <div class="receipt-meta-item"><span class="receipt-meta-label">Date</span><span class="receipt-meta-val">{date_val}</span></div>')
    html.append(f'    <div class="receipt-meta-item"><span class="receipt-meta-label">Time</span><span class="receipt-meta-val">{time_val}</span></div>')
    html.append(f'    <div class="receipt-meta-item"><span class="receipt-meta-label">Currency</span><span class="receipt-meta-val">{currency} ({sym})</span></div>')
    html.append(f'    <div class="receipt-meta-item"><span class="receipt-meta-label">Payment</span><span class="receipt-meta-val">{pay_method}</span></div>')
    html.append(f'  </div>')
    
    # Items Table
    html.append(f'  <table class="receipt-table">')
    html.append(f'    <thead>')
    html.append(f'      <tr>')
    html.append(f'        <th>Item Description</th>')
    html.append(f'        <th class="num-col" style="width: 50px;">Qty</th>')
    html.append(f'        <th class="num-col" style="width: 80px;">Price</th>')
    html.append(f'        <th class="num-col" style="width: 90px;">Total</th>')
    html.append(f'      </tr>')
    html.append(f'    </thead>')
    html.append(f'    <tbody>')
    
    for item in receipt.get("line_items", []):
        qty = item.get("quantity")
        if qty is None:
            qty = 1.0
        try:
            qty_val = float(qty)
            qty_str = f"{qty_val:.0f}" if qty_val.is_integer() else f"{qty_val:.2f}"
        except Exception:
            qty_str = str(qty)
            
        u_price = item.get("unit_price") or 0.0
        t_price = item.get("total_price") or 0.0
        html.append(f'      <tr>')
        html.append(f'        <td>{item.get("name", "Unknown Item")}</td>')
        html.append(f'        <td class="num-col">{qty_str}</td>')
        html.append(f'        <td class="num-col">{sym}{u_price:.2f}</td>')
        html.append(f'        <td class="num-col">{sym}{t_price:.2f}</td>')
        html.append(f'      </tr>')
        
    html.append(f'    </tbody>')
    html.append(f'  </table>')
    
    # Totals Section
    html.append(f'  <div class="receipt-totals-section">')
    if receipt.get("subtotal") is not None:
        html.append(f'    <div class="receipt-total-row"><span class="lbl">Subtotal</span><span class="val">{sym}{receipt["subtotal"]:.2f}</span></div>')
    
    discount_val = receipt.get("discount") or 0.0
    if discount_val > 0:
        html.append(f'    <div class="receipt-total-row discount-row"><span class="lbl">Discount Applied</span><span class="val">-{sym}{discount_val:.2f}</span></div>')
    else:
        html.append(f'    <div class="receipt-total-row"><span class="lbl">Discount</span><span class="val">{sym}0.00</span></div>')
        
    if receipt.get("tax") is not None:
        html.append(f'    <div class="receipt-total-row"><span class="lbl">Tax</span><span class="val">{sym}{receipt["tax"]:.2f}</span></div>')
    if receipt.get("tip") is not None:
        html.append(f'    <div class="receipt-total-row"><span class="lbl">Tip/Gratuity</span><span class="val">{sym}{receipt["tip"]:.2f}</span></div>')
        
    html.append(f'    <div class="receipt-total-row grand-total"><span class="lbl">Total Amount</span><span class="val">{sym}{receipt["total"]:.2f}</span></div>')
    html.append(f'  </div>')
    
    html.append(f'</div>')
    return "\n".join(html)


async def extract_receipt(files, camera_file, method="vision_llm", progress=gr.Progress()) -> tuple[str, str, str, str, str | None]:
    """
    Upload file(s) or webcam capture to the FastAPI backend and return the results.
    Returns a tuple of (formatted_view, status_message, raw_json, validation_info, excel_file_path).
    """
    import traceback
    try:
        from app.schemas import get_currency_symbol

        # Combine input files
        all_files = []
        if files:
            if isinstance(files, list):
                all_files.extend(files)
            else:
                all_files.append(files)
        if camera_file:
            all_files.append(camera_file)

        if not all_files:
            return "", "Please upload a receipt image/PDF/CSV/JSON or capture one using webcam.", "", "", None

        start = time.time()
        successful_ids = []
        results_summary = []
        last_receipt_data = None
        last_validation = None
        all_extracted_data = []

        progress(0, desc="Starting extraction pipeline...")

        for idx, file_item in enumerate(all_files):
            # Resolve actual path and filename robustly
            actual_path = None
            filename = None

            if isinstance(file_item, str):
                actual_path = file_item
                filename = os.path.basename(file_item)
            elif isinstance(file_item, dict):
                actual_path = file_item.get("path") or file_item.get("name")
                filename = file_item.get("orig_name") or (os.path.basename(actual_path) if actual_path else f"upload_{idx+1}")
            else:
                actual_path = getattr(file_item, "path", None) or getattr(file_item, "name", None)
                filename = getattr(file_item, "orig_name", None) or (os.path.basename(actual_path) if actual_path else f"upload_{idx+1}")

            if not actual_path:
                actual_path = str(file_item)
                filename = f"upload_{idx+1}"

            progress((idx / len(all_files)), desc=f"Extracting {filename} ({idx+1}/{len(all_files)})...")

            try:
                # Determine MIME type
                ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
                mime_map = {
                    "jpg": "image/jpeg",
                    "jpeg": "image/jpeg",
                    "png": "image/png",
                    "webp": "image/webp",
                    "pdf": "application/pdf",
                    "csv": "text/csv",
                    "json": "application/json",
                    "txt": "text/plain",
                }
                mime_type = mime_map.get(ext, "application/octet-stream")

                # Read file bytes
                with open(actual_path, "rb") as f:
                    file_bytes = f.read()

                data = None
                # On Render, bypass local network port bindings and invoke the pipeline directly in-process
                if os.getenv("RENDER") == "true":
                    try:
                        from app.main import _process_single_file
                        from fastapi import UploadFile
                        
                        upload_file = UploadFile(
                            file=io.BytesIO(file_bytes),
                            filename=filename,
                            headers={"content-type": mime_type}
                        )
                        result = await _process_single_file(upload_file, method)
                        data = result.model_dump()
                    except Exception as e:
                        print(f"[Render] Direct extraction failed: {e}. Falling back to HTTP.")

                # Fallback (or local development execution path)
                if data is None:
                    with httpx.Client(timeout=120.0) as client:
                        response = client.post(
                            f"{API_BASE}/upload",
                            params={"method": method},
                            files={"file": (filename, file_bytes, mime_type)},
                        )
                    data = response.json()

                if data.get("success"):
                    receipt = data["data"]
                    receipt_id = data.get("id")
                    if receipt_id:
                        successful_ids.append(receipt_id)

                    last_receipt_data = receipt
                    last_validation = data.get("validation")
                    all_extracted_data.append(receipt)

                    sym = get_currency_symbol(receipt.get("currency"))
                    discount_amt = receipt.get("discount") or 0.0
                    results_summary.append({
                        "filename": filename,
                        "vendor": receipt.get("vendor_name", "Unknown"),
                        "date": receipt.get("date", "Unknown"),
                        "total": f"{sym}{receipt.get('total', 0.0):.2f}",
                        "discount": f"-{sym}{discount_amt:.2f}" if discount_amt > 0 else "—",
                        "currency": receipt.get("currency", "INR"),
                        "status": "✓ Success",
                        "confidence": (data.get("validation", {}).get("overall_confidence") or "high").upper()
                    })
                else:
                    error_msg = data.get("error", "Unknown error")
                    results_summary.append({
                        "filename": filename,
                        "vendor": "—",
                        "date": "—",
                        "total": "—",
                        "discount": "—",
                        "currency": "—",
                        "status": f"✗ Failed: {error_msg}",
                        "confidence": "LOW"
                    })

            except Exception as e:
                results_summary.append({
                    "filename": filename,
                    "vendor": "—",
                    "date": "—",
                    "total": "—",
                    "discount": "—",
                    "currency": "—",
                    "status": f"✗ Error: {type(e).__name__}",
                    "confidence": "LOW"
                })

        elapsed = time.time() - start

        if not successful_ids:
            # None of the files succeeded
            summary = "### Batch Processing Failed\n\nNo files were successfully processed."
            status = f"Batch extraction failed in {elapsed:.1f}s."
            return summary, status, "", "", gr.update(visible=False)

        # Generate combined Excel file for download
        excel_path = None
        try:
            from app.export import generate_excel as backend_generate_excel
            excel_bytes = backend_generate_excel(receipt_ids=successful_ids)
            excel_path = os.path.join(tempfile.gettempdir(), f"combined_export_{int(time.time())}.xlsx")
            with open(excel_path, "wb") as f:
                f.write(excel_bytes)
        except Exception as e:
            print(f"Error generating combined Excel: {e}")

        # Build response summary
        if len(all_files) == 1:
            # Single file layout
            receipt = last_receipt_data
            filename = results_summary[0]["filename"] if results_summary else None
            summary = format_single_receipt_html(receipt, filename)
            status = f"Extraction successful in {elapsed:.1f}s"
            editable_json = json.dumps(receipt, indent=2)
            validation_md = _format_validation(last_validation)
        else:
            # Multi-file batch layout
            summary_lines = [
                "## Batch Extraction Summary",
                "",
                f"Successfully processed **{len(successful_ids)}/{len(all_files)}** files.",
                "",
                "| File Name | Vendor | Date | Total | Discount | Currency | Status | Confidence |",
                "| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |"
            ]
            for res in results_summary:
                summary_lines.append(
                    f"| {res['filename']} | {res['vendor']} | {res['date']} | {res['total']} | {res['discount']} | {res['currency']} | {res['status']} | {res['confidence']} |"
                )

            summary_lines.append("")
            summary_lines.append("---")
            summary_lines.append("## Detailed Invoice Breakdowns")
            summary_lines.append("")

            # Append each receipt's HTML card
            for idx, receipt in enumerate(all_extracted_data):
                fname = results_summary[idx]["filename"] if idx < len(results_summary) else None
                summary_lines.append(format_single_receipt_html(receipt, fname))

            summary = "\n".join(summary_lines)
            status = f"Batch extraction complete in {elapsed:.1f}s. Processed {len(successful_ids)}/{len(all_files)} successfully."
            editable_json = json.dumps(all_extracted_data, indent=2)
            validation_md = _format_validation(last_validation)

        excel_update = gr.update(value=excel_path, visible=True) if excel_path else gr.update(visible=False)
        return summary, status, editable_json, validation_md, excel_update

    except Exception as e:
        tb = traceback.format_exc()
        with open("frontend_error.log", "w") as f:
            f.write(tb)
        print(f"CRITICAL ERROR IN GRADIO UI:\n{tb}")
        err_summary = f"### An Error Occurred\n\n**Error Type:** `{type(e).__name__}`\n\n**Message:** {str(e)}\n\nCheck `frontend_error.log` for the full traceback."
        return err_summary, f"Error: {type(e).__name__}", "", "", gr.update(visible=False)


def save_edited_json(json_text: str) -> str:
    """Validate that the edited JSON is still valid."""
    if not json_text.strip():
        return "No data to validate."
    try:
        data = json.loads(json_text)
        return f"JSON is valid. {len(data) if isinstance(data, dict) else len(data)} records/fields."
    except json.JSONDecodeError as e:
        return f"Invalid JSON: {e}"


def export_json(json_text: str):
    """Export the current JSON to a downloadable file."""
    if not json_text.strip():
        return gr.update(visible=False)
    try:
        data = json.loads(json_text)
        
        if isinstance(data, list):
            fname = f"receipts_batch_{int(time.time())}.json"
        else:
            vendor = data.get("vendor_name", "receipt").replace(" ", "_")[:20]
            date = data.get("date", "unknown")
            fname = f"{vendor}_{date}.json"
        
        path = os.path.join(tempfile.gettempdir(), fname)
        with open(path, "w") as f:
            json.dump(data, f, indent=2)
        return gr.update(value=path, visible=True)
    except Exception:
        return gr.update(visible=False)


def export_csv(json_text: str):
    """Export line items as CSV."""
    if not json_text.strip():
        return gr.update(visible=False)
    try:
        data = json.loads(json_text)
        if isinstance(data, list):
            data_list = data
            fname = f"receipts_batch_{int(time.time())}.csv"
        else:
            data_list = [data]
            vendor = data.get("vendor_name", "receipt").replace(" ", "_")[:20]
            date = data.get("date", "unknown")
            fname = f"{vendor}_{date}.csv"
        
        path = os.path.join(tempfile.gettempdir(), fname)
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                "Vendor", "Date", "Currency", "Item Name", "Quantity", "Unit Price", "Total Price",
                "Subtotal", "Discount", "Tax", "Tip", "Total"
            ])
            for receipt in data_list:
                line_items = receipt.get("line_items", [])
                if line_items:
                    for item in line_items:
                        writer.writerow([
                            receipt.get("vendor_name", ""),
                            receipt.get("date", ""),
                            receipt.get("currency", ""),
                            item.get("name", ""),
                            item.get("quantity", ""),
                            item.get("unit_price", ""),
                            item.get("total_price", ""),
                            receipt.get("subtotal", ""),
                            receipt.get("discount", 0.0),
                            receipt.get("tax", ""),
                            receipt.get("tip", ""),
                            receipt.get("total", ""),
                        ])
                else:
                    writer.writerow([
                        receipt.get("vendor_name", ""),
                        receipt.get("date", ""),
                        receipt.get("currency", ""),
                        "", "", "", "",
                        receipt.get("subtotal", ""),
                        receipt.get("discount", 0.0),
                        receipt.get("tax", ""),
                        receipt.get("tip", ""),
                        receipt.get("total", ""),
                    ])
        return gr.update(value=path, visible=True)
    except Exception:
        return gr.update(visible=False)


def export_excel(json_text: str):
    """Export the current receipt data (single or list) to Excel."""
    if not json_text.strip():
        return gr.update(visible=False)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        raw_data = json.loads(json_text)
        if isinstance(raw_data, list):
            data_list = raw_data
        else:
            data_list = [raw_data]

        fname = f"receipts_export_{int(time.time())}.xlsx"
        path = os.path.join(tempfile.gettempdir(), fname)

        wb = Workbook()

        # Sheet 1: Receipts Summary
        ws1 = wb.active
        ws1.title = "Receipts"

        headers1 = [
            "Vendor", "Date", "Currency",
            "Subtotal", "Discount", "Tax", "Tip", "Total",
            "Payment Method"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, item in enumerate(data_list, 2):
            ws1.cell(row=row_idx, column=1, value=item.get("vendor_name", ""))
            ws1.cell(row=row_idx, column=2, value=item.get("date", ""))
            ws1.cell(row=row_idx, column=3, value=item.get("currency", ""))
            ws1.cell(row=row_idx, column=4, value=item.get("subtotal"))
            ws1.cell(row=row_idx, column=5, value=item.get("discount", 0.0))
            ws1.cell(row=row_idx, column=6, value=item.get("tax"))
            ws1.cell(row=row_idx, column=7, value=item.get("tip"))
            ws1.cell(row=row_idx, column=8, value=item.get("total"))
            ws1.cell(row=row_idx, column=9, value=item.get("payment_method", ""))

        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        # Sheet 2: Line Items
        ws2 = wb.create_sheet("Line Items")

        headers2 = [
            "Vendor", "Date", "Item Name", "Quantity", "Unit Price", "Line Total"
        ]

        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row_idx = 2
        for receipt in data_list:
            for item in receipt.get("line_items", []):
                ws2.cell(row=row_idx, column=1, value=receipt.get("vendor_name", ""))
                ws2.cell(row=row_idx, column=2, value=receipt.get("date", ""))
                ws2.cell(row=row_idx, column=3, value=item.get("name", ""))
                ws2.cell(row=row_idx, column=4, value=item.get("quantity"))
                ws2.cell(row=row_idx, column=5, value=item.get("unit_price"))
                ws2.cell(row=row_idx, column=6, value=item.get("total_price"))
                row_idx += 1

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        wb.save(path)
        return gr.update(value=path, visible=True)
    except Exception as e:
        print(f"Error exporting Excel: {e}")
        return gr.update(visible=False)


def _format_validation(validation: dict | None) -> str:
    """Format validation results as readable HTML/markdown."""
    if not validation:
        return "<p style='color: var(--color-muted-ash); font-style: italic;'>No validation data available.</p>"

    lines = []

    # Confidence badge
    confidence = validation.get("overall_confidence", "unknown")
    color_map = {
        "high": ("rgba(16, 185, 129, 0.15)", "#10b981", "HIGH"),
        "medium": ("rgba(245, 158, 11, 0.15)", "#f59e0b", "MEDIUM"),
        "low": ("rgba(239, 68, 68, 0.15)", "#ef4444", "LOW")
    }
    bg, fg, label = color_map.get(confidence, ("rgba(156, 163, 175, 0.15)", "#9ca3af", "UNKNOWN"))
    badge = f'<span style="background-color: {bg}; color: {fg}; padding: 4px 8px; border-radius: 4px; font-weight: 500; font-size: 13px; display: inline-block; letter-spacing: 0.05em; text-transform: uppercase;">{label} CONFIDENCE</span>'
    
    lines.append(f"### Overall Status: {badge}")
    lines.append("")

    # Key metrics
    math_status = "<span style='color: #10b981; font-weight: 500;'>VALID</span>" if validation.get('math_valid') else "<span style='color: #ef4444; font-weight: 500;'>INVALID</span>"
    review_status = "<span style='color: #ef4444; font-weight: 500;'>YES</span>" if validation.get('needs_manual_review') else "<span style='color: #10b981; font-weight: 500;'>NO</span>"
    
    lines.append(f"- **Math Checks:** {math_status}")
    lines.append(f"- **Completeness Score:** {validation.get('completeness_score', 0):.0%}")
    lines.append(f"- **Needs Manual Review:** {review_status}")
    lines.append("")

    # Flags
    flags = validation.get("flags", [])
    if flags:
        lines.append("### Pipeline Flags")
        severity_map = {
            "error": ("rgba(239, 68, 68, 0.15)", "#ef4444", "ERROR"),
            "warning": ("rgba(245, 158, 11, 0.15)", "#f59e0b", "WARN"),
            "info": ("rgba(59, 130, 246, 0.15)", "#3b82f6", "INFO")
        }
        for flag in flags:
            sev = flag.get("severity", "info")
            s_bg, s_fg, s_lbl = severity_map.get(sev, ("rgba(156, 163, 175, 0.15)", "#9ca3af", "INFO"))
            icon = f'<span style="background-color: {s_bg}; color: {s_fg}; padding: 2px 6px; border-radius: 3px; font-size: 11px; font-weight: 500; display: inline-block; margin-right: 6px; letter-spacing: 0.02em;">{s_lbl}</span>'
            field = flag.get("field_name", "")
            msg = flag.get("message", "")
            lines.append(f"- {icon} **{field}**: {msg}")
    else:
        lines.append("<p style='color: #10b981; font-weight: 500; margin-top: 8px;'>✓ No issues found. All validation checks passed.</p>")

    return "\n".join(lines)


# --- Build the Gradio Interface ---

custom_css = """
@import url('https://fonts.googleapis.com/css2?family=Archivo+Black&family=Inter:wght@300;400;450;900&display=swap');

:root {
  --color-cream-paper: #ffedd2;
  --color-void-black: #0d0d0d;
  --color-polished-white: #ffffff;
  --color-hairline-gray: #e5e7eb;
  --color-muted-ash: #9e9e9e;
  --color-surface-charcoal: #1f1f1f;
  
  --font-telka: 'Inter', sans-serif;
  --font-telkaextended: 'Archivo Black', sans-serif;
}

body, html, .gradio-container, .main, .gradio-container > div {
    background-color: var(--color-void-black) !important;
    margin: 0 !important;
    padding: 0 !important;
    font-family: var(--font-telka) !important;
    font-weight: 300 !important;
    max-width: 100% !important;
    border: none !important;
}

#main-container {
    display: flex !important;
    flex-direction: row !important;
    min-height: 100vh !important;
    margin: 0 !important;
    padding: 0 !important;
    gap: 0 !important;
    width: 100% !important;
}

@media (max-width: 768px) {
    #main-container {
        flex-direction: column !important;
    }
    #cream-panel, #void-panel {
        min-height: auto !important;
        height: auto !important;
        padding: 24px !important;
    }
}

#cream-panel {
    background-color: var(--color-cream-paper) !important;
    min-height: 100vh !important;
    padding: 48px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: space-between !important;
    gap: 32px !important;
    border-right: 1px solid var(--color-hairline-gray) !important;
    flex: 1 !important;
}

#void-panel {
    background-color: var(--color-void-black) !important;
    min-height: 100vh !important;
    padding: 48px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: space-between !important;
    gap: 32px !important;
    flex: 1 !important;
    overflow-y: auto !important;
}

/* Typography elements */
.brand-wordmark {
    font-family: var(--font-telka) !important;
    font-weight: 400 !important;
    font-size: 24px !important;
    color: var(--color-void-black) !important;
    text-align: center !important;
    letter-spacing: -0.02em !important;
    margin-bottom: 8px !important;
}

.telka-display {
    font-family: var(--font-telkaextended) !important;
    font-weight: 900 !important;
    font-size: 32px !important;
    line-height: 1.13 !important;
    letter-spacing: 0.01em !important;
    text-transform: uppercase !important;
    color: var(--color-polished-white) !important;
    margin-bottom: 24px !important;
}

.telka-text {
    font-family: var(--font-telka) !important;
    font-weight: 300 !important;
    letter-spacing: -0.02em !important;
    font-size: 16px !important;
    line-height: 1.5 !important;
    color: var(--color-void-black) !important;
}

.tagline {
    font-family: var(--font-telka) !important;
    font-weight: 300 !important;
    font-size: 12px !important;
    color: var(--color-void-black) !important;
    opacity: 0.6 !important;
    text-align: center !important;
    letter-spacing: -0.02em !important;
}

/* Left panel content wrapper styling */
#cream-panel * {
    color: var(--color-void-black) !important;
}

/* Dropzone and video preview visual box */
.video-preview-tile {
    background-color: var(--color-cream-paper) !important;
    border: 1px solid var(--color-hairline-gray) !important;
    border-radius: 10px !important;
    padding: 16px !important;
    box-shadow: none !important;
}

/* Form component overrides for dark panel */
.void-input {
    background-color: var(--color-surface-charcoal) !important;
    border: 1px solid rgba(229, 230, 235, 0.1) !important;
    border-radius: 6px !important;
    padding: 8px !important;
    box-shadow: none !important;
}

.void-input label, .void-input span, .void-input input, .void-input textarea {
    color: var(--color-polished-white) !important;
    font-family: var(--font-telka) !important;
}

.void-input textarea, .void-input input {
    background-color: transparent !important;
    border: none !important;
}

/* Radio button text coloring */
.void-input input[type="radio"]:checked + span {
    color: var(--color-cream-paper) !important;
    font-weight: 450 !important;
}

/* Button variants */
.btn-filled-white {
    background-color: var(--color-polished-white) !important;
    color: var(--color-void-black) !important;
    font-family: var(--font-telka) !important;
    font-weight: 450 !important;
    font-size: 14px !important;
    border-radius: 6px !important;
    border: none !important;
    padding: 12px 16px !important;
    cursor: pointer !important;
    box-shadow: none !important;
    transition: background-color 0.2s ease !important;
    text-align: center !important;
    display: block !important;
    width: 100% !important;
}
.btn-filled-white:hover {
    background-color: #f0f0f0 !important;
}

.btn-outlined-dark {
    background-color: transparent !important;
    color: var(--color-polished-white) !important;
    font-family: var(--font-telka) !important;
    font-weight: 450 !important;
    font-size: 14px !important;
    border-radius: 6px !important;
    border: 1px solid rgba(229, 230, 235, 0.3) !important;
    padding: 12px 16px !important;
    cursor: pointer !important;
    box-shadow: none !important;
    transition: background-color 0.2s ease, border-color 0.2s ease !important;
    text-align: center !important;
    display: block !important;
    width: 100% !important;
}
.btn-outlined-dark:hover {
    background-color: var(--color-surface-charcoal) !important;
    border-color: rgba(229, 230, 235, 0.6) !important;
}

/* Bare tabs */
.void-tabs {
    border: none !important;
    background: transparent !important;
    box-shadow: none !important;
}

.void-tabs .tab-nav {
    border-bottom: 1px solid rgba(229, 230, 235, 0.1) !important;
    background: transparent !important;
    gap: 16px !important;
    box-shadow: none !important;
}

.void-tabs .tab-nav button {
    color: var(--color-muted-ash) !important;
    font-family: var(--font-telka) !important;
    font-weight: 450 !important;
    border: none !important;
    background: transparent !important;
    font-size: 14px !important;
    padding: 8px 0 !important;
    border-radius: 0 !important;
    box-shadow: none !important;
}

.void-tabs .tab-nav button.selected {
    color: var(--color-polished-white) !important;
    border-bottom: 2px solid var(--color-polished-white) !important;
}

.void-tabs .tabitem {
    border: none !important;
    background: transparent !important;
}

.void-tabs .tabitem * {
    color: var(--color-polished-white) !important;
}

/* Footer layout elements */
.footer-row {
    display: flex !important;
    gap: 20px !important;
    justify-content: flex-start !important;
    margin-top: 16px !important;
}

.footer-link {
    color: var(--color-muted-ash) !important;
    font-family: var(--font-telka) !important;
    font-size: 12px !important;
    font-weight: 300 !important;
    text-decoration: none !important;
    transition: color 0.2s ease !important;
}

.footer-link:hover {
    color: var(--color-hairline-gray) !important;
}

.legal-text {
    font-family: var(--font-telka) !important;
    font-size: 12px !important;
    font-weight: 300 !important;
    color: var(--color-muted-ash) !important;
    line-height: 1.5 !important;
    margin-top: 16px !important;
}

.legal-text a {
    color: var(--color-hairline-gray) !important;
    text-decoration: underline !important;
}
"""

# --- Build the Gradio Interface ---

custom_css = """
@import url('https://fonts.googleapis.com/css2?family=Fira+Code&family=Inter:ital,opsz,wght@0,14..32,100..900;1,14..32,100..900&display=swap');

:root {
  /* Colors */
  --color-onyx: #08090a;
  --color-charcoal: #0f1011;
  --color-obsidian: #161718;
  --color-graphite: #23252a;
  --color-iron: #323334;
  --color-steel: #383b3f;
  --color-slate: #62666d;
  --color-fog: #8a8f98;
  --color-mist: #d0d6e0;
  --color-platinum: #e5e5e6;
  --color-snow: #f7f8f8;
  --color-acid-lime: #e4f222;
  --color-indigo: #5e6ad2;
  --color-emerald: #27a644;
  --color-crimson: #eb5757;
  --color-cyan: #02b8cc;

  /* Typography */
  --font-inter: 'Inter', ui-sans-serif, system-ui, -apple-system, sans-serif;
  --font-mono: 'Fira Code', ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace;

  /* Shadows */
  --shadow-sm: rgba(0, 0, 0, 0.4) 0px 2px 4px 0px;
  --shadow-xl: rgba(8, 9, 10, 0.6) 0px 4px 32px 0px;
  --shadow-subtle: rgb(35, 37, 42) 0px 0px 0px 1px inset;
}

footer {
    display: none !important;
}

body, html, .gradio-container, .main, .gradio-container > div {
    background-color: var(--color-onyx) !important;
    margin: 0 !important;
    padding: 0 !important;
    font-family: var(--font-inter) !important;
    font-weight: 300 !important;
    max-width: 100% !important;
    border: none !important;
    color: var(--color-snow) !important;
}

#main-container {
    display: flex !important;
    flex-direction: row !important;
    min-height: 100vh !important;
    margin: 0 !important;
    padding: 0 !important;
    gap: 0 !important;
    width: 100% !important;
}

@media (max-width: 768px) {
    #main-container {
        flex-direction: column !important;
        display: block !important;
    }
    #main-container > div {
        flex-direction: column !important;
        display: flex !important;
        width: 100% !important;
    }
    #left-panel, #right-panel {
        width: 100% !important;
        min-width: 100% !important;
        max-width: 100% !important;
        flex: 1 1 100% !important;
        min-height: auto !important;
        height: auto !important;
        padding: 24px 16px !important;
        border-right: none !important;
        border-bottom: 1px solid var(--color-graphite) !important;
    }
}

@media (max-width: 480px) {
    .receipt-card {
        padding: 16px 12px !important;
    }
    .receipt-meta-grid {
        grid-template-columns: 1fr !important;
        gap: 8px !important;
    }
    .receipt-total-row {
        max-width: 100% !important;
    }
    .receipt-table th, .receipt-table td {
        padding: 8px 2px !important;
        font-size: 11px !important;
    }
}

#left-panel {
    background-color: var(--color-charcoal) !important;
    min-height: 100vh !important;
    padding: 48px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: flex-start !important;
    gap: 24px !important;
    border-right: 1px solid var(--color-graphite) !important;
    flex: 1 !important;
}

#right-panel {
    background-color: var(--color-obsidian) !important;
    min-height: 100vh !important;
    padding: 48px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: space-between !important;
    gap: 32px !important;
    flex: 1 !important;
    overflow-y: auto !important;
}

/* Typography styles */
.brand-title {
    font-family: var(--font-mono) !important;
    font-size: 24px !important;
    font-weight: 500 !important;
    color: var(--color-acid-lime) !important;
    letter-spacing: -0.05em !important;
    margin-bottom: 8px !important;
}

.engine-title {
    font-family: var(--font-inter) !important;
    font-weight: 900 !important;
    font-size: 32px !important;
    letter-spacing: -0.03em !important;
    text-transform: uppercase !important;
    color: var(--color-snow) !important;
    margin-bottom: 24px !important;
}

.sub-text {
    font-size: 14px !important;
    line-height: 1.6 !important;
    color: var(--color-mist) !important;
}

/* Document cards (File Upload & PDF Preview) */
.panel-card {
    background-color: var(--color-onyx) !important;
    border: 1px solid var(--color-graphite) !important;
    border-radius: var(--radius-xl) !important;
    padding: 16px !important;
    box-shadow: var(--shadow-sm) !important;
}

.preview-box {
    background-color: var(--color-onyx) !important;
    border: 1px solid var(--color-graphite) !important;
    border-radius: var(--radius-xl) !important;
    overflow: hidden !important;
    box-shadow: var(--shadow-xl) !important;
    margin-top: 16px !important;
}

/* Inputs on right panel */
.dark-input {
    background-color: var(--color-charcoal) !important;
    border: 1px solid var(--color-graphite) !important;
    border-radius: var(--radius-md) !important;
    padding: 12px !important;
    box-shadow: var(--shadow-md) !important;
}

.dark-input label, .dark-input span, .dark-input input, .dark-input textarea {
    color: var(--color-snow) !important;
    font-family: var(--font-inter) !important;
}

.dark-input textarea, .dark-input input {
    background-color: transparent !important;
    border: none !important;
}

.dark-input input[type="radio"]:checked + span {
    color: var(--color-acid-lime) !important;
    font-weight: 600 !important;
}

/* Buttons */
.btn-primary {
    background-color: var(--color-acid-lime) !important;
    color: var(--color-onyx) !important;
    font-family: var(--font-inter) !important;
    font-weight: 600 !important;
    font-size: 14px !important;
    border-radius: var(--radius-md) !important;
    border: none !important;
    padding: 14px 20px !important;
    cursor: pointer !important;
    transition: all 0.2s ease !important;
    width: 100% !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
}

.btn-primary:hover {
    background-color: #d4e015 !important;
    transform: translateY(-1px) !important;
}

.btn-secondary {
    background-color: transparent !important;
    color: var(--color-snow) !important;
    font-family: var(--font-inter) !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    border-radius: var(--radius-md) !important;
    border: 1px solid var(--color-steel) !important;
    padding: 12px 18px !important;
    cursor: pointer !important;
    transition: all 0.2s ease !important;
    width: 100% !important;
    text-transform: uppercase !important;
    letter-spacing: 0.05em !important;
}

.btn-secondary:hover {
    background-color: var(--color-graphite) !important;
    border-color: var(--color-acid-lime) !important;
    transform: translateY(-1px) !important;
}

/* Dark Tabs */
.dark-tabs {
    border: none !important;
    background: transparent !important;
}

.dark-tabs .tab-nav {
    border-bottom: 1px solid var(--color-graphite) !important;
    background: transparent !important;
    gap: 20px !important;
}

.dark-tabs .tab-nav button {
    color: var(--color-fog) !important;
    font-family: var(--font-inter) !important;
    font-weight: 500 !important;
    border: none !important;
    background: transparent !important;
    font-size: 14px !important;
    padding: 8px 0 !important;
    border-radius: 0 !important;
}

.dark-tabs .tab-nav button:hover {
    color: var(--color-snow) !important;
}

.dark-tabs .tab-nav button.selected {
    color: var(--color-snow) !important;
    border-bottom: 2px solid var(--color-acid-lime) !important;
}

.dark-tabs .tabitem {
    border: none !important;
    background: transparent !important;
    padding: 12px 0 !important;
}

/* Status logs label */
.status-container {
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
    margin-bottom: 6px !important;
}

.status-dot-green {
    width: 6px !important;
    height: 6px !important;
    background-color: var(--color-emerald) !important;
    border-radius: 50% !important;
    display: inline-block !important;
    box-shadow: 0 0 6px var(--color-emerald) !important;
}

.status-label {
    font-size: 12px !important;
    font-weight: 500 !important;
    letter-spacing: 0.05em !important;
    color: var(--color-fog) !important;
}

/* Footer elements */
.footer-row {
    display: flex !important;
    gap: 20px !important;
    justify-content: flex-start !important;
    margin-top: 16px !important;
}

.footer-link {
    color: var(--color-fog) !important;
    font-family: var(--font-inter) !important;
    font-size: 12px !important;
    font-weight: 300 !important;
    text-decoration: none !important;
    transition: color 0.2s ease !important;
}

.footer-link:hover {
    color: var(--color-acid-lime) !important;
}

.legal-text {
    font-family: var(--font-inter) !important;
    font-size: 12px !important;
    font-weight: 300 !important;
    color: var(--color-fog) !important;
    line-height: 1.5 !important;
    margin-top: 16px !important;
}

.legal-text a {
    color: var(--color-mist) !important;
    text-decoration: underline !important;
}

/* Receipt Card Styles */
.receipt-card {
    background-color: var(--color-charcoal) !important;
    border: 1px solid var(--color-graphite) !important;
    border-radius: 8px !important;
    padding: 24px !important;
    margin-bottom: 24px !important;
    box-shadow: var(--shadow-xl) !important;
    font-family: var(--font-inter) !important;
    color: var(--color-snow) !important;
}

.receipt-card-header {
    text-align: center !important;
    margin-bottom: 20px !important;
    border-bottom: 1px dashed var(--color-steel) !important;
    padding-bottom: 16px !important;
}

.receipt-card-title {
    font-size: 22px !important;
    font-weight: 700 !important;
    color: var(--color-acid-lime) !important;
    margin: 0 0 6px 0 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.02em !important;
}

.receipt-card-subtitle {
    font-size: 13px !important;
    color: var(--color-fog) !important;
    margin: 2px 0 !important;
}

.receipt-meta-grid {
    display: grid !important;
    grid-template-columns: repeat(2, 1fr) !important;
    gap: 12px !important;
    font-size: 13px !important;
    margin-bottom: 20px !important;
    background: var(--color-onyx) !important;
    padding: 12px 16px !important;
    border-radius: 6px !important;
    border: 1px solid rgba(229, 230, 235, 0.05) !important;
}

.receipt-meta-item {
    display: flex !important;
    justify-content: space-between !important;
    color: var(--color-snow) !important;
}

.receipt-meta-label {
    color: var(--color-fog) !important;
    font-weight: 500 !important;
}

.receipt-meta-val {
    font-weight: 600 !important;
    text-align: right !important;
}

.receipt-table {
    width: 100% !important;
    border-collapse: collapse !important;
    margin-bottom: 20px !important;
    font-size: 13px !important;
}

.receipt-table th {
    border-bottom: 1px solid var(--color-steel) !important;
    padding: 8px 4px !important;
    text-align: left !important;
    color: var(--color-fog) !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    font-size: 11px !important;
    letter-spacing: 0.05em !important;
}

.receipt-table td {
    padding: 10px 4px !important;
    border-bottom: 1px solid rgba(229, 230, 235, 0.05) !important;
    color: var(--color-snow) !important;
}

.receipt-table th.num-col, .receipt-table td.num-col {
    text-align: right !important;
}

.receipt-totals-section {
    border-top: 1px dashed var(--color-steel) !important;
    padding-top: 16px !important;
    margin-top: 16px !important;
    display: flex !important;
    flex-direction: column !important;
    align-items: flex-end !important;
    gap: 6px !important;
    font-size: 13px !important;
}

.receipt-total-row {
    display: flex !important;
    justify-content: space-between !important;
    width: 100% !important;
    max-width: 260px !important;
}

.receipt-total-row.discount-row {
    color: var(--color-emerald) !important;
    font-weight: 500 !important;
}

.receipt-total-row.grand-total {
    border-top: 1px solid var(--color-steel) !important;
    padding-top: 8px !important;
    margin-top: 4px !important;
    font-size: 16px !important;
    font-weight: 700 !important;
}

.receipt-total-row.grand-total .val {
    color: var(--color-acid-lime) !important;
}

.receipt-total-row .lbl {
    color: var(--color-fog) !important;
}

.receipt-total-row.discount-row .lbl {
    color: var(--color-emerald) !important;
}

.receipt-total-row.grand-total .lbl {
    color: var(--color-snow) !important;
}

.receipt-total-row .val {
    font-weight: 600 !important;
}
"""

with gr.Blocks(
    title="Invoize",
) as demo:
    demo.css = custom_css
    gr.HTML(f"<style>{custom_css}</style>")
    with gr.Row(equal_height=True, elem_id="main-container"):
        # Left Column: Upload panel + Live Image/PDF Preview
        with gr.Column(scale=1, elem_id="left-panel"):
            with gr.Column():
                gr.HTML('<div class="brand-title">Invoize</div>')
                
                with gr.Tabs(elem_classes=["dark-tabs"]):
                    with gr.TabItem("Upload File(s)"):
                        # File Input Area (Multi-file support)
                        file_input = gr.File(
                            label="Drop receipt(s) or invoice(s) here",
                            file_types=[".jpg", ".jpeg", ".png", ".webp", ".pdf", ".csv", ".json"],
                            file_count="multiple",
                            type="filepath",
                            elem_classes=["panel-card"]
                        )
                        # Live Preview Image Component (Visible Uploaded File)
                        preview_image = gr.Image(
                            label="DOCUMENT PREVIEW",
                            interactive=False,
                            elem_classes=["preview-box"]
                        )
                    with gr.TabItem("Camera Capture"):
                        # Webcam input
                        camera_input = gr.Image(
                            label="Capture receipt with webcam",
                            sources=["webcam"],
                            type="filepath",
                            elem_classes=["panel-card"]
                        )
            


        # Right Column: UI action panel + structured output
        with gr.Column(scale=1, elem_id="right-panel"):
            with gr.Column():
                method_input = gr.State("vision_llm")
                
                extract_btn = gr.Button(
                    "EXTRACT STRUCTURED DATA",
                    elem_classes=["btn-primary"]
                )
                
                with gr.Column():
                    gr.HTML(
                        '<div class="status-container">'
                        '<span class="status-dot-green"></span>'
                        '<span class="status-label">PIPELINE STATUS LOGS</span>'
                        '</div>'
                    )
                    status_output = gr.Textbox(
                        show_label=False,
                        value="Ready for document upload.",
                        interactive=False,
                        elem_classes=["dark-input"]
                    )

            # Outputs Tabs
            with gr.Tabs(elem_classes=["dark-tabs"]):
                with gr.TabItem("Formatted View"):
                    formatted_output = gr.Markdown(
                        value="*Upload document to display extracted fields.*"
                    )

                with gr.TabItem("Confidence & Flags"):
                    validation_output = gr.Markdown(
                        value="*Independent math checks and plausibility flags compile here.*"
                    )

                with gr.TabItem("Source JSON"):
                    json_editor = gr.Code(
                        label="SOURCE CODE (EDITABLE)",
                        language="json",
                        lines=15,
                        elem_classes=["dark-input"]
                    )
                    with gr.Row():
                        validate_btn = gr.Button("VALIDATE SCHEMA", elem_classes=["btn-secondary"])
                    validate_output = gr.Textbox(
                        label="SCHEMA VALIDATION RESULTS",
                        interactive=False,
                        elem_classes=["dark-input"]
                    )

            # Downloads & Navigation footer
            with gr.Column():
                with gr.Row():
                    export_json_btn = gr.Button("DOWNLOAD JSON", elem_classes=["btn-secondary"])
                    export_csv_btn = gr.Button("DOWNLOAD CSV", elem_classes=["btn-secondary"])
                    export_excel_btn = gr.Button("DOWNLOAD EXCEL", elem_classes=["btn-secondary"])
                
                json_file = gr.File(label="JSON Download", visible=False, elem_classes=["dark-input"])
                csv_file = gr.File(label="CSV Download", visible=False, elem_classes=["dark-input"])
                excel_file = gr.File(label="Excel Download", visible=False, elem_classes=["dark-input"])

                gr.HTML(
                    '<div class="legal-text">'
                    'By submitting documents, you agree to our '
                    '<a href="#">Terms of Service</a>, '
                    '<a href="#">Acceptable Use Policy</a>, and '
                    '<a href="#">Privacy Policy</a>.'
                    '</div>'
                )

                gr.HTML(
                    '<div class="footer-row">'
                    '<a class="footer-link" href="https://github.com/Natraj16/Invoize" target="_blank">GitHub</a>'
                    '<a class="footer-link" href="https://github.com/Natraj16/Invoize/blob/main/README.md" target="_blank">Docs</a>'
                    '</div>'
                )

    # --- Event Handlers ---
    
    # Auto-generate live preview on file upload/change
    file_input.change(
        fn=preview_file,
        inputs=[file_input],
        outputs=[preview_image],
    )

    extract_btn.click(
        fn=extract_receipt,
        inputs=[file_input, camera_input, method_input],
        outputs=[formatted_output, status_output, json_editor, validation_output, excel_file],
    )

    validate_btn.click(
        fn=save_edited_json,
        inputs=[json_editor],
        outputs=[validate_output],
    )

    export_json_btn.click(
        fn=export_json,
        inputs=[json_editor],
        outputs=[json_file],
    )

    export_csv_btn.click(
        fn=export_csv,
        inputs=[json_editor],
        outputs=[csv_file],
    )

    export_excel_btn.click(
        fn=export_excel,
        inputs=[json_editor],
        outputs=[excel_file],
    )


if __name__ == "__main__":
    port = int(os.getenv("GRADIO_SERVER_PORT", 7860))
    demo.launch(server_port=port)


