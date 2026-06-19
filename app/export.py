"""
Export layer — CSV and Excel generation from stored receipts.

Provides two export formats:
1. CSV: Flat file, one row per line item (with receipt metadata repeated)
   - Universal compatibility (Excel, Google Sheets, pandas, etc.)
   - Easy to import into accounting software

2. Excel: Two-sheet workbook
   - Sheet 1 "Receipts": One row per receipt (summary view)
   - Sheet 2 "Line Items": One row per line item (detail view)
   - Formatted headers and auto-sized columns

Why both formats?
- CSV is programmatic (pipe into scripts, databases, ETL)
- Excel is for humans (accountants, managers, demo reviewers)
"""

import csv
import io
from datetime import datetime
from typing import Optional

from app.storage import get_all_line_items, list_receipts


def generate_csv(receipt_id: Optional[str] = None) -> str:
    """
    Generate CSV content with all receipts and their line items.

    Format: one row per line item with receipt metadata.
    This "flat" format is standard for expense tools — each line is
    independently meaningful.

    Returns CSV as a string.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Receipt ID", "Vendor", "Date", "Currency",
        "Item Name", "Quantity", "Unit Price", "Line Total",
        "Subtotal", "Tax", "Tip", "Total",
        "Confidence", "Needs Review",
    ])

    # Get receipts
    if receipt_id:
        from app.storage import get_receipt
        receipt = get_receipt(receipt_id)
        receipts = [receipt] if receipt else []
    else:
        receipts = list_receipts(limit=10000)

    for receipt in receipts:
        data = receipt.get("extracted_data", {})
        validation = receipt.get("validation_result", {})
        line_items = data.get("line_items", [])

        if line_items:
            for item in line_items:
                writer.writerow([
                    receipt["id"],
                    data.get("vendor_name", ""),
                    data.get("date", ""),
                    data.get("currency", ""),
                    item.get("name", ""),
                    item.get("quantity", ""),
                    item.get("unit_price", ""),
                    item.get("total_price", ""),
                    data.get("subtotal", ""),
                    data.get("tax", ""),
                    data.get("tip", ""),
                    data.get("total", ""),
                    validation.get("overall_confidence", ""),
                    "Yes" if validation.get("needs_manual_review") else "No",
                ])
        else:
            # Receipt with no line items — still include it
            writer.writerow([
                receipt["id"],
                data.get("vendor_name", ""),
                data.get("date", ""),
                data.get("currency", ""),
                "", "", "", "",
                data.get("subtotal", ""),
                data.get("tax", ""),
                data.get("tip", ""),
                data.get("total", ""),
                validation.get("overall_confidence", ""),
                "Yes" if validation.get("needs_manual_review") else "No",
            ])

    return output.getvalue()


def generate_excel(receipt_id: Optional[str] = None) -> bytes:
    """
    Generate an Excel workbook with two sheets.

    Sheet 1 "Receipts": Summary view (one row per receipt)
    Sheet 2 "Line Items": Detail view (one row per item)

    Returns the workbook as bytes (ready to write to response).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()

    # --- Sheet 1: Receipts Summary ---
    ws1 = wb.active
    ws1.title = "Receipts"

    headers1 = [
        "Receipt ID", "Vendor", "Date", "Currency",
        "Subtotal", "Tax", "Tip", "Total",
        "Payment Method", "Confidence", "Needs Review",
        "Uploaded At",
    ]

    # Style the header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Get data
    if receipt_id:
        from app.storage import get_receipt
        receipt = get_receipt(receipt_id)
        receipts = [receipt] if receipt else []
    else:
        receipts = list_receipts(limit=10000)

    for row_idx, receipt in enumerate(receipts, 2):
        data = receipt.get("extracted_data", {})
        validation = receipt.get("validation_result", {})

        ws1.cell(row=row_idx, column=1, value=receipt["id"])
        ws1.cell(row=row_idx, column=2, value=data.get("vendor_name", ""))
        ws1.cell(row=row_idx, column=3, value=data.get("date", ""))
        ws1.cell(row=row_idx, column=4, value=data.get("currency", ""))
        ws1.cell(row=row_idx, column=5, value=data.get("subtotal"))
        ws1.cell(row=row_idx, column=6, value=data.get("tax"))
        ws1.cell(row=row_idx, column=7, value=data.get("tip"))
        ws1.cell(row=row_idx, column=8, value=data.get("total"))
        ws1.cell(row=row_idx, column=9, value=data.get("payment_method", ""))
        ws1.cell(row=row_idx, column=10, value=validation.get("overall_confidence", ""))
        ws1.cell(row=row_idx, column=11, value="Yes" if validation.get("needs_manual_review") else "No")
        ws1.cell(row=row_idx, column=12, value=receipt.get("uploaded_at", ""))

    # Auto-size columns (approximate)
    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # --- Sheet 2: Line Items ---
    ws2 = wb.create_sheet("Line Items")

    headers2 = [
        "Receipt ID", "Vendor", "Date", "Currency",
        "Item Name", "Quantity", "Unit Price", "Line Total",
    ]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for receipt in receipts:
        data = receipt.get("extracted_data", {})
        for item in data.get("line_items", []):
            ws2.cell(row=row_idx, column=1, value=receipt["id"])
            ws2.cell(row=row_idx, column=2, value=data.get("vendor_name", ""))
            ws2.cell(row=row_idx, column=3, value=data.get("date", ""))
            ws2.cell(row=row_idx, column=4, value=data.get("currency", ""))
            ws2.cell(row=row_idx, column=5, value=item.get("name", ""))
            ws2.cell(row=row_idx, column=6, value=item.get("quantity"))
            ws2.cell(row=row_idx, column=7, value=item.get("unit_price"))
            ws2.cell(row=row_idx, column=8, value=item.get("total_price"))
            row_idx += 1

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
